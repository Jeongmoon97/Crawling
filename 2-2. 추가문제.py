# 컨테이너 : “ul.type01 > li”
# 제목 : “a._sp_each_title”
# 언론사 : “span._sp_each_source”
# 1. “키워드 목록＂이라는 두 번째 시트를 만들어 입력한 키워드와 입력시간 저장하기
# 2. 예외 처리와 if 문을 이용하여 “키워드 목록“ 시트에 키워드가 존재하면 “이미 수집된 키워드입니다.”를 출력하며 중단하기

import requests
from bs4 import BeautifulSoup
from datetime import datetime
import openpyxl

# 기존 파일 있을 경우
try:
    wb = openpyxl.load_workbook('keyword.xlsx')
    sheet1 = wb.active
    # wb.active는 첫 번째 시트만을 불러옴. 따라서 sheet2도 선택해줘야 함
    sheet2 = wb['키워드 목록']
    print('불러오기 완료')

# 기존 파일 없을 경우
except:
    wb = openpyxl.Workbook()
    sheet1 = wb.active
    sheet1.title = '키워드 검색 결과'
    sheet2 = wb.create_sheet('키워드 목록')
    sheet1.append(['검색어', '제목', '신문사'])
    sheet2.append(['키워드', '검색 시간'])
    print('새로운 파일을 만들었습니다.')

# 검색 시간
now = datetime.now()
now = now.strftime('%Y-%m-%d %H:%M:%S')

# 키워드 검색
keyword = input('검색어를 입력하세요: ')

# 키워드 리스트업
keyword_list = []
# 시트.rows 를 통해 행들에 엑세스할 수 있음. 각 행을 하나씩 가져오기 위해 for문으로 한 row씩 가져오고 있다.
# 각 row는 그 행 안에 있는 cell 들의 집합으로 여기선 첫 cell 즉 i[0]의 값을 리턴하기 위해 i[0].value 을 사용했다.
for i in sheet2.rows:
    keyword_list.append(i[0].value)
print(keyword_list)

# 새로운 키워드인 경우
if keyword not in keyword_list:
    print(keyword + "에 대해 수집중...")
    for n in range(1, 100, 10):
        raw = requests.get("https://search.naver.com/search.naver?where=news&sm=tab_jum&query=" + keyword + "&start=" + str(n),
                           headers={'User-Agent': 'Morzilla/5.0'})
        html = BeautifulSoup(raw.text, "html.parser")
        articles = html.select("ul.type01>li")

        for ar in articles:
            title = ar.select_one("a._sp_each_title").text.strip().replace(',', '_')
            source = ar.select_one("span._sp_each_source").text.strip().replace(',', '_')
            print(keyword, ':', title, source)
            sheet1.append([keyword, title, source])

    sheet2.append([keyword, now])
    print("수집 완료")

# 시트에 존재하는 키워드인 경우
else:
    print("이미 수집된 키워드입니다.")
    pass

wb.save('keyword.xlsx')